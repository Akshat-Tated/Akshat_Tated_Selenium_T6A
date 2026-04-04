import pytest
import openpyxl

import pytest

from time import sleep

from selenium.webdriver import Chrome,ChromeOptions
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

o= ChromeOptions()
o.add_experimental_option("detach",True)
o.add_argument("--disable-notifications")
driver = Chrome(options=o)

driver.get("https://www.saucedemo.com/")
driver.maximize_window()

driver.implicitly_wait(10)

wb = openpyxl.Workbook()
sheetName = "sheet1"
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws['A1'] ='user'
ws['B1']='password'

"""
standard_user
locked_out_user
problem_user
performance_glitch_user
error_user
visual_user
"""

"""secret_sauce"""

#wb.save('sample.xlsx') #it will save xl file in local directory

ws.append(['standard_user','secret_sauce'])
ws.append(['locked_out_user','secret_sauce'])
ws.append(['problem_user','secret_sauce'])
ws.append(['akshat','tated'])
ws.append(['performance_glitch_user','secret_sauce'])

wb.save('sample.xlsx')

def get_test_data():
    wb = openpyxl.load_workbook('sample.xlsx')
    ws = wb["sheet1"]

    data = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)

    return data

print(get_test_data())

@pytest.mark.parametrize('a,b',get_test_data())
def test_add(a,b):
    u1 =driver.find_element(By.XPATH,'//input[@id="user-name"]')
    u1.send_keys(a)
    u2 = driver.find_element(By.XPATH,'//input[@id="password"]')
    u2.send_keys(b)
    driver.find_element(By.XPATH,'//input[@id="login-button"]').click()
    assert driver.current_url == 'https://www.saucedemo.com/inventory.html', driver.refresh()
    driver.back()
    sleep(5)
    # print(a+b)